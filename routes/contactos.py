from flask import Blueprint, render_template, request, jsonify
from models.contacto import Contacto, ListaContactos
import logging

# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Crear blueprint para contactos
contactos_bp = Blueprint('contactos', __name__)

def manejar_error(error, mensaje_usuario="Error interno del servidor", codigo=500):
    """Función centralizada para manejo de errores"""
    logger.error(f"{mensaje_usuario}: {str(error)}")
    return jsonify({
        'success': False,
        'message': mensaje_usuario,
        'error': str(error) if isinstance(error, ValueError) else None
    }), codigo

@contactos_bp.route('/')
def index():
    """Página principal de contactos"""
    try:
        # Obtener contactos reales desde el almacenamiento local
        contactos = Contacto.obtener_todos()
        estadisticas = Contacto.estadisticas()
        
        return render_template('contactos/contactos.html',
                             pantalla_actual='contactos',
                             contactos=contactos,
                             estadisticas=estadisticas)
        
    except Exception as e:
        logger.error(f"Error en página de contactos: {e}")
        return render_template('contactos/contactos.html',
                             pantalla_actual='contactos',
                             contactos=[],
                             estadisticas={
                                 'total': 0, 
                                 'activos': 0, 
                                 'bloqueados': 0, 
                                 'inactivos': 0, 
                                 'nuevos_7_dias': 0, 
                                 'tasa_activos': 0
                             },
                             error="Error cargando contactos")

@contactos_bp.route('/api', methods=['GET'])
def obtener_contactos():
    """API para obtener todos los contactos"""
    try:
        contactos = Contacto.obtener_todos()
        estadisticas = Contacto.estadisticas()
        
        return jsonify({
            'success': True,
            'data': [contacto.to_dict() for contacto in contactos],
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        return manejar_error(e, "Error obteniendo contactos", 500)

@contactos_bp.route('/api', methods=['POST'])
def crear_contacto():
    """API para crear nuevo contacto"""
    try:
        datos = request.get_json()
        
        if not datos:
            return manejar_error(
                ValueError("No se recibieron datos"),
                "Datos de contacto requeridos",
                400
            )
        
        # Validar datos requeridos
        nombre = datos.get('nombre', '').strip()
        telefono = datos.get('telefono', '').strip()
        
        if not nombre:
            return manejar_error(
                ValueError("Nombre requerido"),
                "El nombre es obligatorio",
                400
            )
        
        if not telefono:
            return manejar_error(
                ValueError("Teléfono requerido"),
                "El teléfono es obligatorio",
                400
            )
        
        # Crear contacto
        contacto = Contacto.crear(
            nombre=nombre,
            telefono=telefono,
            email=datos.get('email'),
            empresa=datos.get('empresa'),
            notas=datos.get('notas'),
            origen='manual'
        )
        
        logger.info(f"Contacto creado: {contacto.nombre} - {contacto.telefono}")
        
        return jsonify({
            'success': True,
            'message': 'Contacto creado exitosamente',
            'data': contacto.to_dict()
        })
        
    except ValueError as ve:
        error_msg = str(ve)
        if "teléfono inválido" in error_msg.lower():
            return manejar_error(ve, f"Número inválido: {error_msg}. El contacto no será guardado.", 400)
        return manejar_error(ve, str(ve), 400)
    except Exception as e:
        return manejar_error(e, "Error creando contacto", 500)

@contactos_bp.route('/api/<contacto_id>', methods=['GET'])
def obtener_contacto(contacto_id):
    """API para obtener un contacto específico"""
    try:
        contacto = Contacto.obtener_por_id(contacto_id)
        
        if not contacto:
            return manejar_error(
                ValueError("Contacto no encontrado"),
                "Contacto no encontrado",
                404
            )
        
        return jsonify({
            'success': True,
            'data': contacto.to_dict()
        })
        
    except Exception as e:
        return manejar_error(e, "Error obteniendo contacto", 500)

@contactos_bp.route('/api/<contacto_id>', methods=['PUT'])
def editar_contacto(contacto_id):
    """API para editar contacto existente"""
    try:
        contacto = Contacto.obtener_por_id(contacto_id)
        
        if not contacto:
            return manejar_error(
                ValueError("Contacto no encontrado"),
                "Contacto no encontrado",
                404
            )
        
        datos = request.get_json()
        
        if not datos:
            return manejar_error(
                ValueError("No se recibieron datos"),
                "Datos de contacto requeridos",
                400
            )
        
        # Validar datos requeridos
        nombre = datos.get('nombre', '').strip()
        telefono = datos.get('telefono', '').strip()
        
        if not nombre:
            return manejar_error(
                ValueError("Nombre requerido"),
                "El nombre es obligatorio",
                400
            )
        
        if not telefono:
            return manejar_error(
                ValueError("Teléfono requerido"),
                "El teléfono es obligatorio",
                400
            )
        
        # Validar teléfono si cambió
        if telefono != contacto.telefono:
            try:
                telefono_limpio = Contacto._limpiar_telefono_estatico(telefono)
                otro_contacto = Contacto.obtener_por_telefono(telefono_limpio)
                if otro_contacto and otro_contacto.id != contacto.id:
                    return manejar_error(
                        ValueError("Teléfono duplicado"),
                        "Ya existe otro contacto con este número de teléfono",
                        400
                    )
                contacto.telefono = telefono_limpio
            except ValueError as ve:
                return manejar_error(ve, f"Número inválido: {str(ve)}. El contacto no será actualizado.", 400)
        
        # Actualizar contacto
        contacto.actualizar(
            nombre=nombre,
            email=datos.get('email'),
            empresa=datos.get('empresa'),
            notas=datos.get('notas')
        )
        
        contacto.guardar()
        
        logger.info(f"Contacto actualizado: {contacto.nombre} - {contacto.telefono}")
        
        return jsonify({
            'success': True,
            'message': 'Contacto actualizado exitosamente',
            'data': contacto.to_dict()
        })
        
    except ValueError as ve:
        return manejar_error(ve, str(ve), 400)
    except Exception as e:
        return manejar_error(e, "Error actualizando contacto", 500)

@contactos_bp.route('/api/<contacto_id>', methods=['DELETE'])
def eliminar_contacto(contacto_id):
    """API para eliminar contacto"""
    try:
        contacto = Contacto.obtener_por_id(contacto_id)
        
        if not contacto:
            return manejar_error(
                ValueError("Contacto no encontrado"),
                "Contacto no encontrado",
                404
            )
        
        nombre_contacto = contacto.nombre
        
        if contacto.eliminar():
            logger.info(f"Contacto eliminado: {nombre_contacto}")
            return jsonify({
                'success': True,
                'message': 'Contacto eliminado exitosamente'
            })
        else:
            return manejar_error(
                Exception("Error en eliminación"),
                "Error eliminando contacto",
                500
            )
        
    except Exception as e:
        return manejar_error(e, "Error eliminando contacto", 500)

@contactos_bp.route('/api/buscar')
def buscar_contactos():
    """API para buscar contactos"""
    try:
        termino = request.args.get('q', '', type=str)
        limite = request.args.get('limite', 50, type=int)
        
        if not termino or not termino.strip():
            return jsonify({'success': True, 'data': []})
        
        contactos = Contacto.buscar(termino)[:limite]
        
        return jsonify({
            'success': True,
            'data': [contacto.to_dict() for contacto in contactos]
        })
        
    except Exception as e:
        return manejar_error(e, "Error en búsqueda", 500)

@contactos_bp.route('/api/estadisticas')
def obtener_estadisticas():
    """API para obtener estadísticas de contactos"""
    try:
        stats = Contacto.estadisticas()
        return jsonify({
            'success': True,
            'data': stats
        })
    except Exception as e:
        return manejar_error(e, "Error obteniendo estadísticas", 500)

# ==================== FUNCIONALIDADES DE EXCEL ====================

@contactos_bp.route('/api/plantilla-excel')
def descargar_plantilla():
    """API para descargar plantilla Excel optimizada"""
    try:
        import pandas as pd
        from io import BytesIO
        from flask import send_file
        
        # Crear DataFrame con solo las columnas esenciales
        datos_ejemplo = {
            'Nombre': [
                'Juan Pérez Gómez',
                'María García López',
                'Carlos López Vásquez'
            ],
            'Telefono': [
                '+593987654321',
                '+593912345678',
                '+593998765432'
            ]
        }
        
        df = pd.DataFrame(datos_ejemplo)
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir datos
            df.to_excel(writer, sheet_name='Contactos', index=False)
            
            # Obtener workbook y worksheet para formateo
            workbook = writer.book
            worksheet = writer.sheets['Contactos']
            
            # Formatear encabezados
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='25D366', end_color='25D366', fill_type='solid')
            
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 25,  # Nombre
                'B': 18   # Telefono
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # IMPORTANTE: Formatear TODA la columna teléfono como texto
            for row in range(1, 1000):
                cell = worksheet.cell(row=row, column=2)
                cell.number_format = '@'
            
            # También aplicar formato específico a los datos existentes
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=2)
                cell.number_format = '@'
                cell.value = str(cell.value)
            
            # Agregar instrucciones en una nueva hoja
            instrucciones_data = {
                'INSTRUCCIONES IMPORTANTES': [
                    '═══════════════════════════════════════════════════════',
                    'COLUMNAS OBLIGATORIAS:',
                    '• Nombre: Nombre completo del contacto',
                    '• Telefono: Formato +593XXXXXXXXX (obligatorio)',
                    '',
                    '🔥 IMPORTANTE - FORMATO DE TELÉFONOS:',
                    '• La columna "Telefono" ya está configurada como TEXTO',
                    '• Al pegar números, asegúrese de que se mantengan como texto',
                    '',
                    'FORMATO DE TELÉFONOS ECUATORIANOS:',
                    '• Debe iniciar con +593',
                    '• Seguido de 9 dígitos que OBLIGATORIAMENTE empiecen con 9',
                    '• Ejemplo CORRECTO: +593987654321',
                    '• Ejemplo INCORRECTO: +593887654321 (no empieza con 9)',
                    '',
                    'VALIDACIONES IMPORTANTES:',
                    '• No se permiten nombres vacíos o solo espacios',
                    '• No se permiten nombres duplicados',
                    '• No se permiten teléfonos duplicados',
                    '• Los teléfonos inválidos serán rechazados automáticamente',
                    '',
                    'FORMATO DEL ARCHIVO:',
                    '• No modifique los nombres de las columnas',
                    '• Máximo 1000 contactos por archivo',
                    '• Formatos soportados: .xlsx, .xls, .csv'
                ]
            }
            
            df_instrucciones = pd.DataFrame(instrucciones_data)
            df_instrucciones.to_excel(writer, sheet_name='INSTRUCCIONES', index=False)
            
            # Formatear hoja de instrucciones
            ws_instrucciones = writer.sheets['INSTRUCCIONES']
            ws_instrucciones.column_dimensions['A'].width = 85
            
            # Formatear texto de instrucciones
            for row in ws_instrucciones.iter_rows():
                for cell in row:
                    if cell.value:
                        if 'OBLIGATORIAS' in str(cell.value):
                            cell.font = Font(bold=True, color='D32F2F', size=12)
                        elif 'IMPORTANTE' in str(cell.value):
                            cell.font = Font(bold=True, color='E53935', size=13)
                        elif str(cell.value).startswith('═'):
                            cell.font = Font(bold=True, color='1976D2', size=14)
                        elif str(cell.value).startswith('•'):
                            cell.font = Font(color='424242', size=10)
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name='plantilla_contactos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except ImportError as ie:
        logger.error(f"Error de importación en plantilla Excel: {ie}")
        return manejar_error(
            ie, 
            f"Falta módulo requerido: {str(ie)}. Instale pandas, numpy y openpyxl", 
            500
        )
    except Exception as e:
        logger.error(f"Error generando plantilla Excel: {e}")
        return manejar_error(e, "Error generando plantilla Excel", 500)

@contactos_bp.route('/api/importar', methods=['POST'])
def importar_contactos():
    """API para importar contactos desde Excel"""
    try:
        if 'archivo' not in request.files:
            return manejar_error(
                ValueError("No se recibió archivo"),
                "Archivo requerido",
                400
            )
        
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return manejar_error(
                ValueError("Archivo vacío"),
                "Seleccione un archivo válido",
                400
            )
        
        # Validar extensión
        extensiones_permitidas = ['.xlsx', '.xls', '.csv']
        extension = archivo.filename.lower()[archivo.filename.rfind('.'):]
        
        if extension not in extensiones_permitidas:
            return manejar_error(
                ValueError("Extensión no válida"),
                "Solo se permiten archivos .xlsx, .xls, .csv",
                400
            )
        
        # Procesar archivo
        from utils.manejador_excel import procesar_archivo_contactos
        resultado = procesar_archivo_contactos(archivo)
        
        # Preparar mensaje de respuesta
        mensaje_base = f'Importación completada: {resultado["procesados"]} de {resultado["total_filas"]} filas procesadas'
        
        alertas = []
        if resultado["duplicados"] > 0:
            alertas.append(f'{resultado["duplicados"]} teléfonos duplicados rechazados')
        if resultado["nombres_repetidos"] > 0:
            alertas.append(f'{resultado["nombres_repetidos"]} nombres duplicados rechazados')
        if resultado["telefonos_invalidos"] > 0:
            alertas.append(f'{resultado["telefonos_invalidos"]} teléfonos inválidos rechazados')
        if resultado["errores"] > 0:
            alertas.append(f'{resultado["errores"]} errores adicionales')
        
        if alertas:
            mensaje_base += f'. ALERTAS: {", ".join(alertas)}'
        
        return jsonify({
            'success': True,
            'message': mensaje_base,
            'data': {
                'estadisticas': {
                    'total_filas': resultado.get('total_filas', 0),
                    'procesados': resultado.get('procesados', 0),
                    'errores': resultado.get('errores', 0),
                    'duplicados': resultado.get('duplicados', 0),
                    'nombres_repetidos': resultado.get('nombres_repetidos', 0),
                    'telefonos_invalidos': resultado.get('telefonos_invalidos', 0),
                    'detalles_errores': resultado.get('detalles_errores', [])
                },
                'contactos': resultado.get('contactos', [])
            }
        })
        
    except ValueError as ve:
        return manejar_error(ve, str(ve), 400)
    except Exception as e:
        logger.error(f"Error importando contactos: {e}")
        return manejar_error(e, "Error importando contactos", 500)